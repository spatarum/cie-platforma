import csv
import io
from datetime import datetime
from typing import Iterable, List

from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from .models import Questionnaire, Submission


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime("%Y-%m-%d %H:%M")


def export_csv(questionnaires: Iterable[Questionnaire]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        [
            "Chestionar",
            "ID chestionar",
            "Termen limită",
            "Capitole",
            "Foi de parcurs",
            "Categorie",
            "Expert",
            "Email",
            "Status",
            "Actualizat la",
            "Trimis la",
            "Nr. întrebare",
            "Întrebare",
            "Răspuns",
        ]
    )

    for q in questionnaires:
        chapters = "; ".join([str(ch) for ch in q.capitole.all().order_by("numar")])
        criteria = "; ".join([c.denumire for c in q.criterii.all().order_by("cod")])
        cat = "General" if getattr(q, "este_general", False) else "Alocat"

        questions = list(q.intrebari.all().order_by("ord"))
        submissions = (
            q.submisii.select_related("expert")
            .prefetch_related("raspunsuri", "raspunsuri__question")
            .filter(status=Submission.STATUS_TRIMIS)
        )

        for sub in submissions:
            expert = sub.expert.get_full_name() or sub.expert.username
            ans_map = {a.question_id: a.text for a in sub.raspunsuri.all()}
            for qu in questions:
                writer.writerow(
                    [
                        q.titlu,
                        q.id,
                        _fmt_dt(q.termen_limita),
                        chapters,
                        criteria,
                        cat,
                        expert,
                        sub.expert.email,
                        dict(sub.STATUS_CHOICES).get(sub.status, sub.status),
                        _fmt_dt(sub.actualizat_la),
                        _fmt_dt(sub.trimis_la),
                        qu.ord,
                        qu.text,
                        ans_map.get(qu.id, ""),
                    ]
                )

    return output.getvalue().encode("utf-8")


def export_xlsx(questionnaires: Iterable[Questionnaire]) -> bytes:
    wb = Workbook()
    # elimină sheet-ul default
    default = wb.active
    wb.remove(default)

    for q in questionnaires:
        title = q.titlu.strip() or f"Chestionar {q.id}"
        sheet_name = title[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{q.id}"
        ws = wb.create_sheet(title=sheet_name)

        questions = list(q.intrebari.all().order_by("ord"))

        headers = [
            "Expert",
            "Email",
            "Telefon",
            "Organizație",
            "Funcție",
            "Categorie",
            "Status",
            "Actualizat la",
            "Trimis la",
        ] + [f"Î{qu.ord}" for qu in questions]

        ws.append(headers)

        submissions = (
            q.submisii.select_related("expert")
            .prefetch_related("raspunsuri", "raspunsuri__question", "expert__profil_expert")
            .filter(status=Submission.STATUS_TRIMIS)
        )

        for sub in submissions:
            expert = sub.expert
            profil = getattr(expert, "profil_expert", None)
            ans_map = {a.question_id: a.text for a in sub.raspunsuri.all()}

            cat = "General" if getattr(q, "este_general", False) else "Alocat"

            row = [
                expert.get_full_name() or expert.username,
                expert.email,
                getattr(profil, "telefon", ""),
                getattr(profil, "organizatie", ""),
                getattr(profil, "functie", ""),
                cat,
                dict(sub.STATUS_CHOICES).get(sub.status, sub.status),
                _fmt_dt(sub.actualizat_la),
                _fmt_dt(sub.trimis_la),
            ]

            for qu in questions:
                row.append(ans_map.get(qu.id, ""))

            ws.append(row)

        # formatare: lățimi
        for idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = 22 if idx <= 9 else 35

        ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_pdf(questionnaires: Iterable[Questionnaire]) -> bytes:
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    margin_x = 40
    y = height - 50

    def write_line(text: str, y_pos: float, font="Helvetica", size=10) -> float:
        c.setFont(font, size)
        # simplu wrap
        max_width = width - 2 * margin_x
        words = text.split()
        line = ""
        for w in words:
            test = (line + " " + w).strip()
            if c.stringWidth(test, font, size) <= max_width:
                line = test
            else:
                c.drawString(margin_x, y_pos, line)
                y_pos -= 14
                line = w
        if line:
            c.drawString(margin_x, y_pos, line)
            y_pos -= 14
        return y_pos

    for q in questionnaires:
        if y < 120:
            c.showPage()
            y = height - 50

        y = write_line(f"Chestionar: {q.titlu}", y, font="Helvetica-Bold", size=12)
        y = write_line(f"Termen limită: {_fmt_dt(q.termen_limita)}", y)
        categorie = "General" if getattr(q, "este_general", False) else "Alocat"
        y = write_line(f"Categorie: {categorie}", y)
        chapters = "; ".join([str(ch) for ch in q.capitole.all().order_by("numar")])
        criteria = "; ".join([c.denumire for c in q.criterii.all().order_by("cod")])
        if chapters:
            y = write_line(f"Capitole: {chapters}", y)
        if criteria:
            y = write_line(f"Foi de parcurs: {criteria}", y)
        y -= 6

        questions = list(q.intrebari.all().order_by("ord"))
        submissions = (
            q.submisii.select_related("expert")
            .prefetch_related("raspunsuri", "raspunsuri__question")
            .filter(status=Submission.STATUS_TRIMIS)
        )

        if not submissions:
            y = write_line("(Nu există răspunsuri încă)", y)
            y -= 10
            continue

        for sub in submissions:
            if y < 120:
                c.showPage()
                y = height - 50

            expert = sub.expert.get_full_name() or sub.expert.username
            y = write_line(f"Expert: {expert} ({sub.expert.email})", y, font="Helvetica-Bold")
            y = write_line(f"Status: {dict(sub.STATUS_CHOICES).get(sub.status, sub.status)} | Actualizat: {_fmt_dt(sub.actualizat_la)}", y)
            ans_map = {a.question_id: a.text for a in sub.raspunsuri.all()}

            for qu in questions:
                if y < 120:
                    c.showPage()
                    y = height - 50
                y = write_line(f"Î{qu.ord}. {qu.text}", y, font="Helvetica-Bold")
                rasp = ans_map.get(qu.id, "") or "(fără răspuns)"
                y = write_line(f"Răspuns: {rasp}", y)
                y -= 4

            y -= 10

        y -= 10

    c.save()
    return buffer.getvalue()
