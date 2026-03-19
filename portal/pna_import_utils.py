from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from django.contrib.auth.models import User
from django.db import transaction

from .models import Chapter, Criterion, EUAct, PnaInstitution, PnaProject, PnaProjectEUAct

_TEMPLATE_MAIN_SHEET = "Proiecte_PNA"
_TEMPLATE_ACTS_SHEET = "Acte_UE"
_TEMPLATE_INFO_SHEET = "Instructiuni"
_TEMPLATE_LISTS_SHEET = "Liste"


_RO_MONTHS = {
    "ianuarie": 1,
    "februarie": 2,
    "martie": 3,
    "aprilie": 4,
    "mai": 5,
    "iunie": 6,
    "iulie": 7,
    "august": 8,
    "septembrie": 9,
    "octombrie": 10,
    "noiembrie": 11,
    "decembrie": 12,
}


_TEMPLATE_PROJECT_COLUMNS = [
    (
        "Cod unic identificabil",
        22,
        "Recomandat pentru update. Dacă este completat, importul caută proiectul mai întâi după acest cod.",
    ),
    ("Nr. acțiune", 14, "Identificatorul acțiunii din PNA. Folosit ca fallback la identificare împreună cu capitolul/foaia."),
    ("Denumire proiect", 48, "Obligatoriu. Titlul proiectului / acțiunii normative."),
    ("Descriere", 42, "Descriere scurtă a proiectului."),
    ("Cluster PNA", 24, "Clusterul PNA, dacă este cazul."),
    ("Capitol (număr)", 16, "Completează fie numărul capitolului (ex. 1), fie codul foii de parcurs."),
    ("Capitol (denumire)", 30, "Opțional. Dacă lipsește capitolul în sistem, poate fi creat cu această denumire."),
    ("Foaie de parcurs (cod)", 22, "Completează fie codul foii de parcurs (ex. CR, PAR, RoL), fie capitolul."),
    ("Foaie de parcurs (denumire)", 32, "Opțional. Dacă lipsește foaia de parcurs în sistem, poate fi creată cu această denumire."),
    ("Status implementare", 30, "Alege una dintre etapele disponibile din lista de validare."),
    ("Instituția principală", 36, "Instituția principală responsabilă. Dacă nu există în listă, va fi creată automat."),
    (
        "Alte instituții responsabile",
        42,
        "Listează instituțiile separate prin ; (punct și virgulă). Instituțiile noi sunt create automat.",
    ),
    ("Contact responsabil", 28, "Nume și funcție persoană responsabilă."),
    ("Email contact", 24, "Emailul persoanei responsabile."),
    ("Termen aprobare în Guvern", 22, "Lună/an. Poți introduce 2026-10, Octombrie 2026 sau o dată Excel (ziua va fi ignorată)."),
    ("Termen aprobare în Parlament", 24, "Lună/an. Poți introduce 2026-12, Decembrie 2026 sau o dată Excel (ziua va fi ignorată)."),
    ("Termen actualizat aprobare în Guvern", 28, "Lună/an. Dacă este completat, acesta este termenul efectiv pentru Guvern."),
    ("Consultări publice în Parlament", 24, "Dată calendaristică (zi/lună/an)."),
    ("Intrare planificată în vigoare", 28, "Text liber: ex. Ianuarie 2026 / T2 2027 etc."),
    ("Complexitate", 18, "Valori acceptate: 1-5 sau eticheta completă (ex. 3 - Medie)."),
    ("Prioritate (1-3)", 18, "Valori acceptate: 1-3 sau eticheta completă."),
    ("Disponibilitate expertiză internă", 28, "Valori acceptate: 1-3 sau eticheta completă."),
    ("Volum de muncă (zile)", 18, "Număr întreg de zile de lucru estimate."),
    ("Necesită expertiză externă", 20, "Da / Nu."),
    ("Disponibilitate expertiză externă", 34, "Cine livrează expertiza externă / detalii."),
    ("Parteneri societate civilă", 34, "Organizații / parteneri care trebuie consultați."),
    ("Cost 2026 (mii lei)", 18, "Valoare numerică."),
    ("Cost 2027 (mii lei)", 18, "Valoare numerică."),
    ("Cost 2028 (mii lei)", 18, "Valoare numerică."),
    ("Cost 2029 (mii lei)", 18, "Valoare numerică."),
    ("Riscuri", 34, "Riscuri / blocaje / observații."),
    ("Raport de extindere 2023", 20, "Da / Nu."),
    ("Raport de extindere 2024", 20, "Da / Nu."),
    ("Raport de extindere 2025", 20, "Da / Nu."),
    ("Raport de extindere 2026", 20, "Da / Nu."),
    ("Raport de extindere 2027", 20, "Da / Nu."),
    ("Planul de creștere economică", 24, "Da / Nu."),
    ("Necesită avizare Comisia Europeană", 28, "Da / Nu."),
    ("Comentariu PNA", 30, "Comentariu din monitorizarea PNA."),
    ("Întârziat 2025", 16, "Da / Nu."),
    ("Note explicative", 28, "Note explicative / context."),
    ("Partener de dezvoltare", 28, "Partener de dezvoltare, dacă există."),
    ("Executor acțiune", 28, "Persoana / echipa care execută acțiunea."),
    ("Cost total (mii lei)", 18, "Cost total estimativ în mii lei."),
    ("Acoperit din bugetul de stat (mii lei)", 24, "Valoare numerică."),
    ("Acoperit din asistență externă (mii lei)", 28, "Valoare numerică."),
    ("Costuri neacoperite (mii lei)", 24, "Valoare numerică."),
    (
        "Acte normative în vigoare de transpunere",
        42,
        "Dacă există mai multe acte, separă-le prin rând nou sau prin ;.",
    ),
    ("Prioritate PNA (text)", 22, "Textul de prioritate din sursa PNA, dacă vrei să îl păstrezi separat."),
]

_TEMPLATE_ACT_COLUMNS = [
    (
        "Cod unic identificabil proiect",
        24,
        "Recomandat. Dacă este completat, actul UE se atașează proiectului identificat prin acest cod.",
    ),
    ("Nr. acțiune proiect", 18, "Fallback de identificare a proiectului împreună cu capitolul / foaia."),
    ("Denumire proiect", 42, "Fallback de identificare a proiectului împreună cu capitolul / foaia."),
    ("Capitol proiect (număr)", 18, "Completează doar dacă proiectul nu este identificat prin cod unic."),
    ("Foaie de parcurs proiect (cod)", 22, "Completează doar dacă proiectul nu este identificat prin cod unic."),
    ("CELEX sau link", 30, "Obligatoriu. Acceptă cod CELEX (ex. 32014L0041) sau link EUR-Lex."),
    ("Denumire act UE", 44, "Denumirea actului UE."),
    ("Tip act UE", 18, "Directivă / Regulament / Recomandare etc."),
    ("Link act UE", 34, "Link EUR-Lex, dacă vrei să îl salvezi explicit."),
    ("Tip transpunere", 20, "Opțional: Total / Parțial."),
]


def _strip_accents(value: Any) -> str:
    text = str(value or "")
    return "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")


def _norm_header(value: Any) -> str:
    text = _strip_accents(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _norm_text(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s+", " ", text)
    return text


def _norm_inst_name(value: Any) -> str:
    return _norm_text(value).lower()


def _parse_chapter_from_label(label: Any) -> int | None:
    s = str(label or "").strip()
    if not s:
        return None
    m = re.search(r"(capitolul|cap\.?|chapter)?\s*(\d{1,2})", s, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(2))
    except Exception:
        return None


def _parse_primary_criterion_code(raw: Any) -> str | None:
    s = str(raw or "").strip()
    if not s:
        return None
    s_up = _strip_accents(s).upper()
    for code in ["ROL", "PAR", "FDI", "GP", "CR"]:
        if re.search(rf"\b{code}\b", s_up):
            return "RoL" if code == "ROL" else code
    token = re.split(r"[\s,;|/–\-]+", s_up, maxsplit=1)[0].strip()
    if not token:
        return None
    return "RoL" if token == "ROL" else token[:20]


def _extract_celex_from_link_or_code(raw: Any) -> tuple[str, str]:
    s = str(raw or "").strip()
    if not s:
        return "", ""
    url = s if s.lower().startswith("http") else ""
    m = re.search(r"CELEX:([0-9A-Za-z]+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip(), url
    m = re.search(r"uri=CELEX:([0-9A-Za-z]+)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip(), url
    celex = s.replace("CELEX:", "").replace("celex:", "").strip()
    celex = re.sub(r"[^0-9A-Za-z]", "", celex)
    return celex, url


def _to_date_from_month_value(raw_value: Any, fallback_year: int | None = None) -> date | None:
    if raw_value in (None, ""):
        return None

    if isinstance(raw_value, datetime):
        return raw_value.date().replace(day=1)
    if isinstance(raw_value, date):
        return raw_value.replace(day=1)

    s = str(raw_value).strip()
    if not s:
        return None

    for fmt in ["%Y-%m", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%Y", "%m-%Y"]:
        try:
            d = datetime.strptime(s, fmt).date()
            return d.replace(day=1)
        except Exception:
            pass

    m = re.match(r"^(\d{4})[./](\d{1,2})$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)

    m = re.match(r"^([A-Za-zăâîșțĂÂÎȘȚ]+)\s+(\d{4})$", s)
    if m:
        mon = _strip_accents(m.group(1)).lower()
        year = int(m.group(2))
        if mon in _RO_MONTHS:
            return date(year, _RO_MONTHS[mon], 1)

    mon_only = _strip_accents(s).lower()
    if mon_only in _RO_MONTHS and fallback_year:
        return date(int(fallback_year), _RO_MONTHS[mon_only], 1)

    try:
        num = float(s)
        dt = openpyxl.utils.datetime.from_excel(num)
        if isinstance(dt, datetime):
            return dt.date().replace(day=1)
        if isinstance(dt, date):
            return dt.replace(day=1)
    except Exception:
        pass

    return None


def _to_date_value(raw_value: Any) -> date | None:
    """Parsează o valoare de tip dată (zi/lună/an).

    Acceptă:
    - date / datetime
    - string-uri (YYYY-MM-DD, DD.MM.YYYY, etc.)
    - valori Excel serial (număr)

    Dacă se primește doar lună/an, setează ziua = 1.
    """

    if raw_value in (None, ""):
        return None

    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value

    s = str(raw_value).strip()
    if not s:
        return None

    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m", "%m/%Y", "%m-%Y"]:
        try:
            d = datetime.strptime(s, fmt).date()
            # dacă nu există zi în format (ex. %Y-%m), ziua va fi 1
            return d
        except Exception:
            pass

    # Formate gen 2026.10 / 2026/10
    m = re.match(r"^(\d{4})[./](\d{1,2})$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)

    try:
        num = float(s)
        dt = openpyxl.utils.datetime.from_excel(num)
        if isinstance(dt, datetime):
            return dt.date()
        if isinstance(dt, date):
            return dt
    except Exception:
        pass

    return None


_TRUE_VALUES = {"da", "true", "1", "x", "yes", "y", "ok"}
_FALSE_VALUES = {"nu", "false", "0", "no", "n", ""}


def _to_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return bool(value)
    s = _strip_accents(str(value)).strip().lower()
    if s in _TRUE_VALUES:
        return True
    if s in _FALSE_VALUES:
        return False
    return default


def _to_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _choice_int(value: Any, labels: list[tuple[int, str]]) -> int | None:
    if value in (None, ""):
        return None
    num = _to_int(value)
    valid = {k for k, _ in labels}
    if num in valid:
        return num
    norm = _norm_header(value)
    for key, label in labels:
        if norm == _norm_header(label) or norm == _norm_header(f"{key} {label}"):
            return key
    return None


def _status_code(value: Any) -> str | None:
    if value in (None, ""):
        return None
    raw = str(value).strip()
    if not raw:
        return None

    norm = _norm_header(raw)

    # match direct code / labels din enum-ul curent
    for code, label in PnaProject.STATUS_IMPLEMENTARE_CHOICES:
        if raw == code or norm == _norm_header(label) or norm == _norm_header(f"{code} {label}"):
            return code

    # fallback: acceptă și statusuri vechi / sinonime (pentru compatibilitate la import)
    alt = {
        _norm_header("Neînceput"): PnaProject.STATUS_NEINITIAT,
        _norm_header("IN_LUCRU_GUVERN"): PnaProject.STATUS_INITIAT_GUVERN,
        _norm_header("IN_AVIZARE_GUVERN"): PnaProject.STATUS_AVIZARE_GUVERN,
        _norm_header("ADOPTAT_GUVERN"): PnaProject.STATUS_INITIAT_PARLAMENT,
        _norm_header("IN_AVIZARE_CE"): PnaProject.STATUS_COORDONARE_CE,
        _norm_header("IN_PROCEDURA_PARLAMENT"): PnaProject.STATUS_AVIZARE_PARLAMENT,
        _norm_header("ADOPTAT_PARLAMENT"): PnaProject.STATUS_ADOPTAT_FINAL,
        _norm_header("Neinițiat"): PnaProject.STATUS_NEINITIAT,
        _norm_header("În lucru la Guvern"): PnaProject.STATUS_INITIAT_GUVERN,
        _norm_header("Inițiat în Guvern"): PnaProject.STATUS_INITIAT_GUVERN,
        _norm_header("În avizare la Guvern"): PnaProject.STATUS_AVIZARE_GUVERN,
        _norm_header("În coordonare cu Comisia Europeană"): PnaProject.STATUS_COORDONARE_CE,
        _norm_header("În avizare la Comisia Europeană"): PnaProject.STATUS_COORDONARE_CE,
        _norm_header("În aprobare la Guvern"): PnaProject.STATUS_APROBARE_GUVERN,
        _norm_header("Adoptat de Guvern"): PnaProject.STATUS_INITIAT_PARLAMENT,
        _norm_header("Inițiat în Parlament"): PnaProject.STATUS_INITIAT_PARLAMENT,
        _norm_header("În avizare la Parlament"): PnaProject.STATUS_AVIZARE_PARLAMENT,
        _norm_header("În procedură legislativă la Parlament"): PnaProject.STATUS_AVIZARE_PARLAMENT,
        _norm_header("Adoptat în prima lectură"): PnaProject.STATUS_ADOPTAT_PRIMA_LECTURA,
        _norm_header("Adoptat în lectura finală de Parlament"): PnaProject.STATUS_ADOPTAT_FINAL,
        _norm_header("Adoptat de Parlament"): PnaProject.STATUS_ADOPTAT_FINAL,
    }
    return alt.get(norm)


def _tip_transpunere(value: Any) -> str:
    if value in (None, ""):
        return ""
    s = _strip_accents(str(value)).strip().lower()
    if "total" in s:
        return PnaProjectEUAct.TIP_TRANSPUNERE_TOTAL
    if "par" in s:
        return PnaProjectEUAct.TIP_TRANSPUNERE_PARTIAL
    if s in {PnaProjectEUAct.TIP_TRANSPUNERE_TOTAL.lower(), PnaProjectEUAct.TIP_TRANSPUNERE_PARTIAL.lower()}:
        return s.upper()
    return ""


def _split_multi_values(raw: Any) -> list[str]:
    if raw in (None, ""):
        return []
    parts = re.split(r"[;\n\r|,/]+", str(raw))
    return [re.sub(r"\s+", " ", p).strip() for p in parts if str(p).strip()]


def _parse_report_years(raw: Any) -> dict[int, bool]:
    text = _strip_accents(raw).lower()
    years = {2023: False, 2024: False, 2025: False, 2026: False, 2027: False}
    if not text:
        return years
    for y in years:
        if str(y) in text:
            years[y] = True
    return years


def _header_alias_map() -> dict[str, str]:
    aliases: dict[str, list[str]] = {
        "pna_cod_unic": ["Cod unic identificabil", "COD UNIC IDENTIFICABIL", "cod_unic_identificabil", "Cod unic identificabil proiect"],
        "pna_nr_actiune": ["Nr. acțiune", "NR. D/O ACȚIUNE", "nr_actiune", "nr do actiune", "Nr. acțiune proiect"],
        "titlu": ["Denumire proiect", "Titlu", "ACȚIUNE NORMATIVĂ", "actiune normativa"],
        "descriere": ["Descriere"],
        "pna_cluster": ["Cluster PNA", "Cluster", "cluster_pna"],
        "capitol_numar": ["Capitol (număr)", "Capitol", "capitol_numar", "Capitol proiect (număr)"],
        "capitol_denumire": ["Capitol (denumire)", "capitol_denumire"],
        "foaie_cod": ["Foaie de parcurs (cod)", "FOAIE DE PARCURS", "foaie_cod", "foaie de parcurs", "Foaie de parcurs proiect (cod)"],
        "foaie_denumire": ["Foaie de parcurs (denumire)", "foaie_denumire"],
        "status_implementare": ["Status implementare", "status_implementare"],
        "institutie_principala": ["Instituția principală", "INSTITUȚIA RESPONSABILĂ", "institutia responsabila"],
        "institutii_responsabile": [
            "Alte instituții responsabile",
            "INSTITUȚIA CO-RESPONSABILĂ",
            "institutia co-responsabila",
            "institutii_responsabile",
        ],
        "contact_responsabil": ["Contact responsabil", "contact_responsabil"],
        "contact_responsabil_email": ["Email contact", "contact_responsabil_email"],
        "termen_aprobare_guvern": ["Termen aprobare în Guvern", "TERMEN APROBARE ÎN GUVERN", "termen_aprobare_guvern"],
        "termen_aprobare_parlament": ["Termen aprobare în Parlament", "LUNĂ", "termen_aprobare_parlament"],
        "termen_actualizat_aprobare_guvern": [
            "Termen actualizat aprobare în Guvern",
            "termen_actualizat_aprobare_guvern",
        ],
        "consultari_publice_parlament": ["Consultări publice în Parlament", "consultari_publice_parlament"],
        "intrare_planificata_vigoare": ["Intrare planificată în vigoare", "intrare_planificata_vigoare"],
        "complexitate": ["Complexitate", "complexitate"],
        "prioritate": ["Prioritate (1-3)", "prioritate"],
        "expertiza_interna": ["Disponibilitate expertiză internă", "expertiza_interna"],
        "volum_munca_zile": ["Volum de muncă (zile)", "volum_munca_zile"],
        "necesita_expertiza_externa": ["Necesită expertiză externă", "necesita_expertiza_externa"],
        "disponibilitate_expertiza_externa": [
            "Disponibilitate expertiză externă",
            "disponibilitate_expertiza_externa",
        ],
        "parteneri_societate_civila": ["Parteneri societate civilă", "parteneri_societate_civila"],
        "cost_2026": ["Cost 2026 (mii lei)", "cost_2026", "cost 2026"],
        "cost_2027": ["Cost 2027 (mii lei)", "cost_2027", "cost 2027"],
        "cost_2028": ["Cost 2028 (mii lei)", "cost_2028", "cost 2028"],
        "cost_2029": ["Cost 2029 (mii lei)", "cost_2029", "cost 2029"],
        "riscuri": ["Riscuri", "riscuri"],
        "raport_extindere_2023": ["Raport de extindere 2023", "raport_extindere_2023"],
        "raport_extindere_2024": ["Raport de extindere 2024", "raport_extindere_2024"],
        "raport_extindere_2025": ["Raport de extindere 2025", "raport_extindere_2025"],
        "raport_extindere_2026": ["Raport de extindere 2026", "raport_extindere_2026"],
        "raport_extindere_2027": ["Raport de extindere 2027", "raport_extindere_2027"],
        "plan_crestere_economica": ["Planul de creștere economică", "plan_crestere_economica"],
        "necesita_avizare_comisia_europeana": [
            "Necesită avizare Comisia Europeană",
            "Este necesară avizarea Comisiei Europene",
            "necesita_avizare_comisia_europeana",
        ],
        "comentariu_pna": ["Comentariu PNA", "Comentariu", "COMENTARIU", "comentariu_pna"],
        "intarziat_2025": ["Întârziat 2025", "intarziat_2025"],
        "note_explicative": ["Note explicative", "note_explicative"],
        "partener_de_dezvoltare": ["Partener de dezvoltare", "PARTENERUL DE DEZVOLTARE", "partener_de_dezvoltare"],
        "executor_actiune": ["Executor acțiune", "EXECUTOR ACȚIUNE", "EXECUTOR ACȚIUNE (2)", "executor_actiune"],
        "cost_total_mii_lei": ["Cost total (mii lei)", "COSTURI ESTIMATIVE (mii lei)", "cost_total_mii_lei"],
        "cost_buget_stat_mii_lei": [
            "Acoperit din bugetul de stat (mii lei)",
            "ACOPERIT DIN BUGETUL DE STAT (mii lei)",
            "cost_buget_stat_mii_lei",
        ],
        "cost_asistenta_externa_mii_lei": [
            "Acoperit din asistență externă (mii lei)",
            "ACOPERIT DIN ASISTENȚĂ EXTERNĂ (mii lei)",
            "cost_asistenta_externa_mii_lei",
        ],
        "cost_neacoperite_mii_lei": [
            "Costuri neacoperite (mii lei)",
            "COSTURI NEACOPERITE (mii lei)",
            "cost_neacoperite_mii_lei",
        ],
        "acte_normative_transpunere_existente": [
            "Acte normative în vigoare de transpunere",
            "Acte normative în vigoare de transpunere a actului UE",
        ],
        "pna_prioritate_text": ["Prioritate PNA (text)", "Prioritate", "pna_prioritate_text"],
        "raport_extindere_text": ["RAPORT EXTINDERE", "raport_extindere"],
        "anul_adoptarii": ["ANUL ADOPTĂRII", "anul adoptarii"],
        "celex": ["CELEX", "CELEX sau link", "celex", "celex sau link"],
        "denumire_act_ue": ["DENUMIRE ACT UE", "Denumire act UE", "denumire_act_ue"],
        "tip_act_ue": ["Tip document UE", "Tip act UE", "tip_act_ue"],
        "link_act_ue": ["Link act UE", "link_act_ue"],
        "tip_transpunere": [
            "Tip transpunere",
            "Intentia de transpunere a actului UE prin actiunea normativa din coloana I (se va marca celula cu coloare verde in cazul transpunerii totale, cu galben - pentru transpunere partiala)",
        ],
    }
    out: dict[str, str] = {}
    for canonical, vals in aliases.items():
        for v in vals:
            out[_norm_header(v)] = canonical
    return out


_HEADER_ALIAS_MAP = _header_alias_map()


def _header_index(headers: list[Any]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, header in enumerate(headers):
        canonical = _HEADER_ALIAS_MAP.get(_norm_header(header))
        if canonical and canonical not in idx:
            idx[canonical] = i
    return idx


def _collect_institution_columns(headers: list[Any]) -> dict[int, str]:
    reserved = set(_HEADER_ALIAS_MAP.values())
    out: dict[int, str] = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        normalized = _norm_header(header)
        if normalized in _HEADER_ALIAS_MAP:
            continue
        text = str(header).strip()
        if not text:
            continue
        # Considerăm drept coloane de instituții pe cele care par denumiri de instituții.
        if " - " in text or " – " in text or text.startswith("Agen") or text.startswith("Minister") or text.startswith("Parlament"):
            out[i] = text
    return out


def _make_inst_resolver() -> tuple[dict[str, PnaInstitution], callable]:
    cache = {_norm_inst_name(i.nume): i for i in PnaInstitution.objects.all()}

    def get_inst(name: str) -> PnaInstitution | None:
        name0 = _norm_text(name)
        if not name0:
            return None
        key = _norm_inst_name(name0)
        obj = cache.get(key)
        if obj:
            return obj
        obj = PnaInstitution.objects.create(nume=name0[:400])
        cache[key] = obj
        return obj

    return cache, get_inst


def _scope_key(chapter: Chapter | None, criterion: Criterion | None) -> tuple[str, str]:
    if chapter:
        return ("chapter", str(chapter.numar))
    if criterion:
        return ("criterion", criterion.cod)
    return ("none", "")


def _resolve_scope_from_values(
    *,
    capitol_numar: Any = None,
    capitol_denumire: Any = None,
    foaie_cod: Any = None,
    foaie_denumire: Any = None,
    existing: PnaProject | None = None,
    prefer_chapter_when_both: bool = False,
) -> tuple[Chapter | None, Criterion | None, str | None]:
    ch_num = _parse_chapter_from_label(capitol_numar)
    if ch_num is None and capitol_numar not in (None, ""):
        ch_num = _to_int(capitol_numar)

    criterion_code = _parse_primary_criterion_code(foaie_cod)

    if ch_num and criterion_code:
        if not prefer_chapter_when_both:
            return None, None, "Completează fie capitolul, fie foaia de parcurs, nu ambele."
        criterion_code = None

    if not ch_num and not criterion_code:
        if existing:
            return existing.chapter, existing.criterion, None
        return None, None, "Lipsește capitolul / foaia de parcurs."

    if ch_num:
        chapter = Chapter.objects.filter(numar=ch_num).first()
        if not chapter:
            den = _norm_text(capitol_denumire) or f"Capitol {ch_num}"
            chapter = Chapter.objects.create(numar=ch_num, denumire=den[:255])
        return chapter, None, None

    criterion = Criterion.objects.filter(cod__iexact=criterion_code).first()
    if not criterion:
        den = _norm_text(foaie_denumire) or criterion_code
        criterion = Criterion.objects.create(cod=criterion_code[:20], denumire=den[:255])
    return None, criterion, None


class _ImportContext:
    def __init__(self) -> None:
        self.created_ids: set[int] = set()
        self.updated_ids: set[int] = set()
        self.error_count = 0
        self.report_rows: list[tuple[str, str, str, str]] = []
        self.project_cache: dict[tuple[str, str, str | None], PnaProject] = {}

    def report(self, row_ref: str, identifier: str, status: str, message: str) -> None:
        self.report_rows.append((row_ref, identifier, status, message))
        if status == "ERROR":
            self.error_count += 1

    def cache_project(self, obj: PnaProject) -> None:
        scope_type, scope_val = _scope_key(obj.chapter, obj.criterion)
        if obj.pna_cod_unic:
            self.project_cache[("code", _norm_text(obj.pna_cod_unic).lower(), None)] = obj
        if obj.pna_nr_actiune:
            self.project_cache[("nr", _norm_text(obj.pna_nr_actiune).lower(), f"{scope_type}:{scope_val}")] = obj
        self.project_cache[("title", _norm_text(obj.titlu).lower(), f"{scope_type}:{scope_val}")] = obj

    def mark_created(self, obj: PnaProject) -> None:
        self.created_ids.add(obj.id)
        self.cache_project(obj)

    def mark_updated(self, obj: PnaProject) -> None:
        if obj.id not in self.created_ids:
            self.updated_ids.add(obj.id)
        self.cache_project(obj)


def _lookup_existing_project(
    *,
    code: str = "",
    nr_actiune: str = "",
    titlu: str = "",
    chapter: Chapter | None = None,
    criterion: Criterion | None = None,
    cache: dict[tuple[str, str, str | None], PnaProject] | None = None,
) -> PnaProject | None:
    scope_type, scope_val = _scope_key(chapter, criterion)
    scope_key = f"{scope_type}:{scope_val}" if scope_type != "none" else None

    if code:
        key = ("code", _norm_text(code).lower(), None)
        if cache and key in cache:
            return cache[key]
        obj = PnaProject.objects.filter(pna_cod_unic__iexact=code).first()
        if obj:
            return obj

    if nr_actiune and scope_key:
        key = ("nr", _norm_text(nr_actiune).lower(), scope_key)
        if cache and key in cache:
            return cache[key]
        obj = PnaProject.objects.filter(pna_nr_actiune__iexact=nr_actiune, chapter=chapter, criterion=criterion).first()
        if obj:
            return obj

    if titlu and scope_key:
        key = ("title", _norm_text(titlu).lower(), scope_key)
        if cache and key in cache:
            return cache[key]
        obj = PnaProject.objects.filter(titlu__iexact=titlu, chapter=chapter, criterion=criterion).first()
        if obj:
            return obj

    return None


def _set_if_changed(obj: Any, field: str, value: Any) -> bool:
    current = getattr(obj, field)
    if current != value:
        setattr(obj, field, value)
        return True
    return False


def _upsert_project(
    data: dict[str, Any],
    *,
    user: User,
    ctx: _ImportContext,
    clear_missing: bool,
    get_inst,
) -> tuple[PnaProject | None, str, str]:
    title = _norm_text(data.get("titlu"))
    code = _norm_text(data.get("pna_cod_unic"))
    nr_actiune = _norm_text(data.get("pna_nr_actiune"))

    # mai întâi căutăm prin cod, apoi rezolvăm scope-ul dacă există în fișier
    existing_by_code = _lookup_existing_project(code=code, cache=ctx.project_cache) if code else None

    chapter, criterion, scope_error = _resolve_scope_from_values(
        capitol_numar=data.get("capitol_numar"),
        capitol_denumire=data.get("capitol_denumire"),
        foaie_cod=data.get("foaie_cod"),
        foaie_denumire=data.get("foaie_denumire"),
        existing=existing_by_code,
    )
    if scope_error:
        return None, "ERROR", scope_error

    existing = existing_by_code or _lookup_existing_project(
        code=code,
        nr_actiune=nr_actiune,
        titlu=title,
        chapter=chapter,
        criterion=criterion,
        cache=ctx.project_cache,
    )

    create_new = existing is None
    obj = existing or PnaProject(creat_de=user)
    changed = False

    # required / identifiers
    if not title and create_new:
        return None, "ERROR", "Lipsește denumirea proiectului."
    if title:
        changed = _set_if_changed(obj, "titlu", title) or changed
    if chapter or criterion:
        changed = _set_if_changed(obj, "chapter", chapter) or changed
        changed = _set_if_changed(obj, "criterion", criterion) or changed

    field_values = {
        "pna_cod_unic": code or "",
        "pna_nr_actiune": nr_actiune or "",
        "descriere": _norm_text(data.get("descriere")) if data.get("descriere") is not None else ("" if clear_missing else None),
        "pna_cluster": _norm_text(data.get("pna_cluster")) if data.get("pna_cluster") is not None else ("" if clear_missing else None),
        "status_implementare": _status_code(data.get("status_implementare")) if data.get("status_implementare") not in (None, "") else (PnaProject.STATUS_NEINITIAT if create_new else None),
        "contact_responsabil": _norm_text(data.get("contact_responsabil")) if data.get("contact_responsabil") is not None else ("" if clear_missing else None),
        "contact_responsabil_email": _norm_text(data.get("contact_responsabil_email")) if data.get("contact_responsabil_email") is not None else ("" if clear_missing else None),
        "termen_aprobare_guvern": _to_date_from_month_value(data.get("termen_aprobare_guvern"), fallback_year=_to_int(data.get("anul_adoptarii"))),
        "termen_aprobare_parlament": _to_date_from_month_value(data.get("termen_aprobare_parlament"), fallback_year=_to_int(data.get("anul_adoptarii"))),
        "termen_actualizat_aprobare_guvern": _to_date_from_month_value(data.get("termen_actualizat_aprobare_guvern"), fallback_year=_to_int(data.get("anul_adoptarii"))),
        "consultari_publice_parlament": _to_date_value(data.get("consultari_publice_parlament")),
        "intrare_planificata_vigoare": _norm_text(data.get("intrare_planificata_vigoare")) if data.get("intrare_planificata_vigoare") is not None else ("" if clear_missing else None),
        "complexitate": _choice_int(data.get("complexitate"), PnaProject.COMPLEXITATE_CHOICES),
        "prioritate": _choice_int(data.get("prioritate"), PnaProject.PRIORITATE_CHOICES),
        "expertiza_interna": _choice_int(data.get("expertiza_interna"), PnaProject.EXPERTIZA_INTERNA_CHOICES),
        "volum_munca_zile": _to_int(data.get("volum_munca_zile")),
        "necesita_expertiza_externa": _to_bool(data.get("necesita_expertiza_externa"), default=False) if (clear_missing or data.get("necesita_expertiza_externa") is not None) else None,
        "disponibilitate_expertiza_externa": _norm_text(data.get("disponibilitate_expertiza_externa")) if data.get("disponibilitate_expertiza_externa") is not None else ("" if clear_missing else None),
        "parteneri_societate_civila": _norm_text(data.get("parteneri_societate_civila")) if data.get("parteneri_societate_civila") is not None else ("" if clear_missing else None),
        "cost_2026": _to_decimal(data.get("cost_2026")),
        "cost_2027": _to_decimal(data.get("cost_2027")),
        "cost_2028": _to_decimal(data.get("cost_2028")),
        "cost_2029": _to_decimal(data.get("cost_2029")),
        "riscuri": _norm_text(data.get("riscuri")) if data.get("riscuri") is not None else ("" if clear_missing else None),
        "raport_extindere_2023": _to_bool(data.get("raport_extindere_2023"), default=False) if (clear_missing or data.get("raport_extindere_2023") is not None) else None,
        "raport_extindere_2024": _to_bool(data.get("raport_extindere_2024"), default=False) if (clear_missing or data.get("raport_extindere_2024") is not None) else None,
        "raport_extindere_2025": _to_bool(data.get("raport_extindere_2025"), default=False) if (clear_missing or data.get("raport_extindere_2025") is not None) else None,
        "raport_extindere_2026": _to_bool(data.get("raport_extindere_2026"), default=False) if (clear_missing or data.get("raport_extindere_2026") is not None) else None,
        "raport_extindere_2027": _to_bool(data.get("raport_extindere_2027"), default=False) if (clear_missing or data.get("raport_extindere_2027") is not None) else None,
        "plan_crestere_economica": _to_bool(data.get("plan_crestere_economica"), default=False) if (clear_missing or data.get("plan_crestere_economica") is not None) else None,
        "necesita_avizare_comisia_europeana": _to_bool(data.get("necesita_avizare_comisia_europeana"), default=False) if (clear_missing or data.get("necesita_avizare_comisia_europeana") is not None) else None,
        "comentariu_pna": _norm_text(data.get("comentariu_pna")) if data.get("comentariu_pna") is not None else ("" if clear_missing else None),
        "intarziat_2025": _to_bool(data.get("intarziat_2025"), default=False) if (clear_missing or data.get("intarziat_2025") is not None) else None,
        "note_explicative": _norm_text(data.get("note_explicative")) if data.get("note_explicative") is not None else ("" if clear_missing else None),
        "partener_de_dezvoltare": _norm_text(data.get("partener_de_dezvoltare")) if data.get("partener_de_dezvoltare") is not None else ("" if clear_missing else None),
        "executor_actiune": _norm_text(data.get("executor_actiune")) if data.get("executor_actiune") is not None else ("" if clear_missing else None),
        "cost_total_mii_lei": _to_decimal(data.get("cost_total_mii_lei")),
        "cost_buget_stat_mii_lei": _to_decimal(data.get("cost_buget_stat_mii_lei")),
        "cost_asistenta_externa_mii_lei": _to_decimal(data.get("cost_asistenta_externa_mii_lei")),
        "cost_neacoperite_mii_lei": _to_decimal(data.get("cost_neacoperite_mii_lei")),
        "acte_normative_transpunere_existente": _norm_text(data.get("acte_normative_transpunere_existente")) if data.get("acte_normative_transpunere_existente") is not None else ("" if clear_missing else None),
        "pna_prioritate_text": _norm_text(data.get("pna_prioritate_text")) if data.get("pna_prioritate_text") is not None else ("" if clear_missing else None),
    }

    # extra booleans din coloana RAPORT EXTINDERE (fișierul PNA original)
    if data.get("raport_extindere_text") not in (None, ""):
        rep = _parse_report_years(data.get("raport_extindere_text"))
        field_values.update(
            {
                "raport_extindere_2023": rep[2023],
                "raport_extindere_2024": rep[2024],
                "raport_extindere_2025": rep[2025],
                "raport_extindere_2026": rep[2026],
                "raport_extindere_2027": rep[2027],
            }
        )

    for field, value in field_values.items():
        if value is None and not clear_missing:
            continue
        changed = _set_if_changed(obj, field, value) or changed

    # instituții
    principal_raw = data.get("institutie_principala")
    other_raw = data.get("institutii_responsabile")
    principal_obj = get_inst(principal_raw) if (principal_raw not in (None, "") or clear_missing) else obj.institutie_principala_ref
    other_values = []
    if other_raw not in (None, ""):
        other_values.extend(_split_multi_values(other_raw))
    if clear_missing and other_raw in (None, ""):
        other_values = []

    other_objs = []
    seen = set()
    for nm in other_values:
        inst = get_inst(nm)
        if not inst:
            continue
        if principal_obj and inst.id == principal_obj.id:
            continue
        key = _norm_inst_name(inst.nume)
        if key in seen:
            continue
        seen.add(key)
        other_objs.append(inst)

    if principal_raw not in (None, "") or clear_missing:
        changed = _set_if_changed(obj, "institutie_principala_ref", principal_obj) or changed
        changed = _set_if_changed(obj, "institutie_principala", (principal_obj.nume if principal_obj else "")[:300]) or changed

    if clear_missing or other_raw is not None:
        other_txt = ", ".join([o.nume for o in other_objs])[:300]
        changed = _set_if_changed(obj, "institutie_coreponsabila", other_txt) or changed

    if obj.arhivat:
        obj.arhivat = False
        obj.arhivat_la = None
        changed = True

    obj.full_clean(exclude=["acte_ue", "institutii_responsabile"])
    obj.save()

    m2m_changed = False
    if clear_missing or other_raw is not None:
        old_ids = sorted(obj.institutii_responsabile.values_list("id", flat=True))
        new_ids = sorted([o.id for o in other_objs])
        if old_ids != new_ids:
            obj.institutii_responsabile.set(other_objs)
            m2m_changed = True

    if create_new:
        ctx.mark_created(obj)
        return obj, "CREATED", "Creat"
    if changed or m2m_changed:
        ctx.mark_updated(obj)
        return obj, "UPDATED", "Actualizat"

    ctx.cache_project(obj)
    return obj, "OK", "Neschimbat"


def _attach_eu_act(
    *,
    project: PnaProject,
    celex_or_link: Any,
    denumire: Any,
    tip_document: Any,
    url_value: Any,
    tip_transpunere_value: Any,
) -> tuple[str, str]:
    celex, url_from_celex = _extract_celex_from_link_or_code(celex_or_link)
    url = str(url_value or "").strip() or url_from_celex
    if not celex:
        return "ERROR", "CELEX / link invalid"

    den = _norm_text(denumire) or celex
    tip_doc = _norm_text(tip_document)

    act, created = EUAct.objects.get_or_create(
        celex=celex,
        defaults={"denumire": den, "tip_document": tip_doc, "url": url},
    )

    changed_act = False
    if den and act.denumire != den:
        act.denumire = den
        changed_act = True
    if tip_doc and act.tip_document != tip_doc:
        act.tip_document = tip_doc
        changed_act = True
    if url and act.url != url:
        act.url = url
        changed_act = True
    if changed_act:
        act.save()

    link, created_link = PnaProjectEUAct.objects.get_or_create(project=project, eu_act=act)
    tip_trans = _tip_transpunere(tip_transpunere_value)
    if tip_trans and link.tip_transpunere != tip_trans:
        link.tip_transpunere = tip_trans
        link.save(update_fields=["tip_transpunere"])
        return "UPDATED", "Act UE actualizat / atașat"
    if created or created_link:
        return "UPDATED", "Act UE atașat"
    return "OK", "Act UE deja atașat"


def _template_sheet(wb, *candidates: str):
    normalized = {_norm_header(nm): nm for nm in wb.sheetnames}
    for cand in candidates:
        if cand in wb.sheetnames:
            return wb[cand]
        nm = normalized.get(_norm_header(cand))
        if nm:
            return wb[nm]
    return None


def _import_template_workbook(wb, *, user: User, ctx: _ImportContext) -> None:
    sheet = _template_sheet(wb, _TEMPLATE_MAIN_SHEET, "Proiecte PNA")
    if sheet is None:
        raise ValueError("Nu am găsit sheet-ul «Proiecte_PNA» în template.")

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h for h in header_row]
    idx = _header_index(headers)
    if "titlu" not in idx:
        raise ValueError("Template invalid: lipsește coloana «Denumire proiect».")

    _, get_inst = _make_inst_resolver()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in row):
            continue
        data = {key: row[col] if col < len(row) else None for key, col in idx.items()}
        identifier = _norm_text(data.get("titlu")) or _norm_text(data.get("pna_cod_unic")) or "(fără titlu)"
        try:
            with transaction.atomic():
                _, status, message = _upsert_project(data, user=user, ctx=ctx, clear_missing=True, get_inst=get_inst)
            ctx.report(f"{sheet.title}!{row_idx}", identifier, status, message)
        except Exception as exc:
            ctx.report(f"{sheet.title}!{row_idx}", identifier, "ERROR", str(exc))

    # sheet acte UE (opțional)
    acts_sheet = _template_sheet(wb, _TEMPLATE_ACTS_SHEET, "Acte UE")
    if acts_sheet is None:
        return

    act_header_row = next(acts_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    act_headers = [h for h in act_header_row]
    act_idx = _header_index(act_headers)
    if "celex" not in act_idx:
        return

    for row_idx, row in enumerate(acts_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in row):
            continue
        data = {key: row[col] if col < len(row) else None for key, col in act_idx.items()}
        identifier = _norm_text(data.get("celex")) or "(act UE)"
        try:
            chapter, criterion, scope_error = _resolve_scope_from_values(
                capitol_numar=data.get("capitol_numar"),
                foaie_cod=data.get("foaie_cod"),
            )
            if scope_error and not _norm_text(data.get("pna_cod_unic")):
                raise ValueError(scope_error)

            project = _lookup_existing_project(
                code=_norm_text(data.get("pna_cod_unic")),
                nr_actiune=_norm_text(data.get("pna_nr_actiune")),
                titlu=_norm_text(data.get("titlu")),
                chapter=chapter,
                criterion=criterion,
                cache=ctx.project_cache,
            )
            if not project:
                raise ValueError("Nu am găsit proiectul pentru acest act UE. Completează codul unic sau denumirea + capitol/foaie.")

            with transaction.atomic():
                status, message = _attach_eu_act(
                    project=project,
                    celex_or_link=data.get("celex"),
                    denumire=data.get("denumire_act_ue"),
                    tip_document=data.get("tip_act_ue"),
                    url_value=data.get("link_act_ue"),
                    tip_transpunere_value=data.get("tip_transpunere"),
                )
            if status == "UPDATED":
                ctx.mark_updated(project)
            else:
                ctx.cache_project(project)
            ctx.report(f"{acts_sheet.title}!{row_idx}", identifier, status, message)
        except Exception as exc:
            ctx.report(f"{acts_sheet.title}!{row_idx}", identifier, "ERROR", str(exc))


def _import_source_pna_workbook(wb, *, user: User, ctx: _ImportContext) -> None:
    sheet = _template_sheet(wb, "Acțiuni_PNA", "Actiuni_PNA")
    if sheet is None:
        for nm in wb.sheetnames:
            if "pna" in _norm_header(nm):
                sheet = wb[nm]
                break
    if sheet is None:
        raise ValueError("Nu am găsit sheet-ul «Acțiuni_PNA» în fișier.")

    headers = [h for h in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = _header_index(headers)
    if "titlu" not in idx:
        raise ValueError("Fișier invalid: lipsește coloana «ACȚIUNE NORMATIVĂ».")

    inst_columns = _collect_institution_columns(headers)
    _, get_inst = _make_inst_resolver()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in row):
            continue
        data = {key: row[col] if col < len(row) else None for key, col in idx.items()}
        identifier = _norm_text(data.get("titlu")) or f"rând {row_idx}"

        # scope: sursa originală are etichete, nu numere/coduri separate
        if data.get("capitol_numar") not in (None, ""):
            data["capitol_denumire"] = data.get("capitol_numar")
        if data.get("foaie_cod") not in (None, ""):
            data["foaie_denumire"] = data.get("foaie_cod")

        # instituții din coloanele marcate cu Da/X/1
        flagged_institutions = []
        for col_idx, inst_name in inst_columns.items():
            val = row[col_idx] if col_idx < len(row) else None
            if _to_bool(val, default=False):
                flagged_institutions.append(inst_name)

        principal_parts = _split_multi_values(data.get("institutie_principala"))
        principal_name = principal_parts[0] if principal_parts else _norm_text(data.get("institutie_principala"))
        other_names = []
        if len(principal_parts) > 1:
            other_names.extend(principal_parts[1:])
        other_names.extend(_split_multi_values(data.get("institutii_responsabile")))
        other_names.extend(flagged_institutions)

        # dedupe + scoatem instituția principală din lista secundară
        unique_others = []
        seen = set()
        for nm in other_names:
            key = _norm_inst_name(nm)
            if not key or key == _norm_inst_name(principal_name) or key in seen:
                continue
            seen.add(key)
            unique_others.append(nm)

        data["institutie_principala"] = principal_name
        data["institutii_responsabile"] = "; ".join(unique_others)

        # executor acțiune poate fi în două coloane distincte -> le concatenăm
        exec_values = []
        for header_name in ["EXECUTOR ACȚIUNE", "EXECUTOR ACȚIUNE (2)"]:
            try:
                col = headers.index(header_name)
            except ValueError:
                col = None
            if col is not None and col < len(row) and row[col] not in (None, ""):
                exec_values.append(str(row[col]).strip())
        if exec_values:
            data["executor_actiune"] = "\n\n".join([v for v in exec_values if v])

        # sursa originală are luna de Parlament + anul adoptării
        data.setdefault("termen_aprobare_parlament", data.get("termen_aprobare_parlament"))

        # unele coloane pot apărea repetat pentru actele normative existente
        transp_existing = []
        for i, h in enumerate(headers):
            hh = _norm_header(h)
            if hh.startswith(_norm_header("Acte normative în vigoare de transpunere a actului UE")):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    txt = str(val).strip()
                    if txt and txt not in transp_existing:
                        transp_existing.append(txt)
        if transp_existing:
            data["acte_normative_transpunere_existente"] = "\n".join(transp_existing)

        try:
            # Fișierul sursă PNA poate avea atât capitol, cât și foaie de parcurs completate pe același rând.
            # În sistem, proiectul rămâne atașat unui singur scope, iar regula existentă este să preferăm capitolul.
            if data.get("capitol_numar") not in (None, "") and data.get("foaie_cod") not in (None, ""):
                data["foaie_cod"] = None
                data["foaie_denumire"] = None
            with transaction.atomic():
                project, status, message = _upsert_project(data, user=user, ctx=ctx, clear_missing=False, get_inst=get_inst)
            ctx.report(f"{sheet.title}!{row_idx}", identifier, status, message)
            if not project:
                continue

            # act UE din rând (dacă există)
            if data.get("celex") not in (None, ""):
                with transaction.atomic():
                    act_status, act_message = _attach_eu_act(
                        project=project,
                        celex_or_link=data.get("celex"),
                        denumire=data.get("denumire_act_ue"),
                        tip_document=data.get("tip_act_ue"),
                        url_value=data.get("link_act_ue"),
                        tip_transpunere_value=data.get("tip_transpunere"),
                    )
                if act_status == "UPDATED":
                    ctx.mark_updated(project)
                ctx.report(f"{sheet.title}!{row_idx}", f"{identifier} / act UE", act_status, act_message)
        except Exception as exc:
            ctx.report(f"{sheet.title}!{row_idx}", identifier, "ERROR", str(exc))


def run_pna_import_workbook(wb, *, user: User) -> dict[str, Any]:
    ctx = _ImportContext()
    if _template_sheet(wb, _TEMPLATE_MAIN_SHEET, "Proiecte PNA") is not None:
        mode = "template"
        _import_template_workbook(wb, user=user, ctx=ctx)
    else:
        mode = "pna_source"
        _import_source_pna_workbook(wb, user=user, ctx=ctx)

    return {
        "mode": mode,
        "nr_create": len(ctx.created_ids),
        "nr_update": len(ctx.updated_ids),
        "nr_error": ctx.error_count,
        "report_rows": ctx.report_rows,
    }


def _apply_header_style(ws, total_columns: int) -> None:
    fill = PatternFill("solid", fgColor="0B3D91")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D5D9E2")
    for col_idx in range(1, total_columns + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_columns)}1"
    ws.row_dimensions[1].height = 34


def _add_comment(cell, text: str) -> None:
    cell.comment = Comment(text, "CIE")


def _add_list_validation(ws, cell_range: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_pna_import_template_workbook() -> Workbook:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = _TEMPLATE_INFO_SHEET

    title_font = Font(size=16, bold=True, color="0B3D91")
    subtitle_font = Font(size=11, bold=True)
    ws_info["A1"] = "Template import PNA"
    ws_info["A1"].font = title_font
    ws_info["A3"] = "Cum se folosește"
    ws_info["A3"].font = subtitle_font
    instructions = [
        "1. Completează câte un rând per proiect în sheet-ul Proiecte_PNA.",
        "2. Pentru mai multe acte UE la același proiect folosește sheet-ul Acte_UE (un rând per act).",
        "3. Pentru termene completează luna/an (ex. 2026-10, Octombrie 2026 sau o dată Excel setată pe prima zi a lunii).",
        "4. Pentru instituții multiple folosește separatorul ; în coloana «Alte instituții responsabile». Instituțiile noi vor fi create automat la import.",
        "5. Regula de identificare la update este: Cod unic identificabil → Nr. acțiune + capitol/foaie → Denumire proiect + capitol/foaie.",
        "6. Importul actualizează sau creează proiecte și atașează acte UE. Nu șterge actele UE existente care nu apar în fișier.",
    ]
    for idx, line in enumerate(instructions, start=4):
        ws_info[f"A{idx}"] = line
        ws_info[f"A{idx}"].alignment = Alignment(wrap_text=True)
    ws_info["A12"] = "Sheet-uri incluse"
    ws_info["A12"].font = subtitle_font
    ws_info["A13"] = f"• {_TEMPLATE_MAIN_SHEET}: datele principale ale proiectelor"
    ws_info["A14"] = f"• {_TEMPLATE_ACTS_SHEET}: acte UE relevante (mai multe rânduri per proiect)"
    ws_info["A15"] = f"• {_TEMPLATE_LISTS_SHEET}: liste de referință pentru valori și validări"
    ws_info.column_dimensions["A"].width = 120

    ws_main = wb.create_sheet(_TEMPLATE_MAIN_SHEET)
    for col_idx, (header, width, comment) in enumerate(_TEMPLATE_PROJECT_COLUMNS, start=1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        _add_comment(cell, comment)
        ws_main.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_header_style(ws_main, len(_TEMPLATE_PROJECT_COLUMNS))

    # formatare coloane lună / costuri pentru primele 500 de rânduri
    month_headers = {
        "Termen aprobare în Guvern",
        "Termen aprobare în Parlament",
        "Termen actualizat aprobare în Guvern",
    }
    date_headers = {
        "Consultări publice în Parlament",
    }

    money_headers = {
        "Cost 2026 (mii lei)",
        "Cost 2027 (mii lei)",
        "Cost 2028 (mii lei)",
        "Cost 2029 (mii lei)",
        "Cost total (mii lei)",
        "Acoperit din bugetul de stat (mii lei)",
        "Acoperit din asistență externă (mii lei)",
        "Costuri neacoperite (mii lei)",
    }
    for col_idx, (header, _width, _comment) in enumerate(_TEMPLATE_PROJECT_COLUMNS, start=1):
        rng = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}500"
        if header in month_headers:
            for row_no in range(2, 501):
                ws_main[f"{get_column_letter(col_idx)}{row_no}"].number_format = "mmmm yyyy"
        elif header in date_headers:
            for row_no in range(2, 501):
                ws_main[f"{get_column_letter(col_idx)}{row_no}"].number_format = "dd.mm.yyyy"
        elif header in money_headers:
            for row_no in range(2, 501):
                ws_main[f"{get_column_letter(col_idx)}{row_no}"].number_format = "#,##0.00"

    ws_acts = wb.create_sheet(_TEMPLATE_ACTS_SHEET)
    for col_idx, (header, width, comment) in enumerate(_TEMPLATE_ACT_COLUMNS, start=1):
        cell = ws_acts.cell(row=1, column=col_idx, value=header)
        _add_comment(cell, comment)
        ws_acts.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_header_style(ws_acts, len(_TEMPLATE_ACT_COLUMNS))

    ws_lists = wb.create_sheet(_TEMPLATE_LISTS_SHEET)
    ws_lists["A1"] = "Da / Nu"
    ws_lists["B1"] = "Status implementare"
    ws_lists["C1"] = "Complexitate"
    ws_lists["D1"] = "Prioritate"
    ws_lists["E1"] = "Expertiză internă"
    ws_lists["F1"] = "Tip transpunere"
    ws_lists["G1"] = "Foi de parcurs"
    ws_lists["H1"] = "Capitole"
    ws_lists["I1"] = "Instituții"

    for i, v in enumerate(["Da", "Nu"], start=2):
        ws_lists[f"A{i}"] = v
    for i, (_code, label) in enumerate(PnaProject.STATUS_IMPLEMENTARE_CHOICES, start=2):
        ws_lists[f"B{i}"] = label
    for i, (key, label) in enumerate(PnaProject.COMPLEXITATE_CHOICES, start=2):
        ws_lists[f"C{i}"] = f"{key} - {label}"
    for i, (key, label) in enumerate(PnaProject.PRIORITATE_CHOICES, start=2):
        ws_lists[f"D{i}"] = f"{key} - {label}"
    for i, (key, label) in enumerate(PnaProject.EXPERTIZA_INTERNA_CHOICES, start=2):
        ws_lists[f"E{i}"] = f"{key} - {label}"
    for i, val in enumerate(["Total", "Parțial"], start=2):
        ws_lists[f"F{i}"] = val

    crit_row = 2
    for c in Criterion.objects.order_by("cod"):
        ws_lists[f"G{crit_row}"] = c.cod
        ws_lists[f"G{crit_row}"].comment = Comment(c.denumire, "CIE")
        crit_row += 1

    ch_row = 2
    for ch in Chapter.objects.order_by("numar"):
        ws_lists[f"H{ch_row}"] = f"{ch.numar} - {ch.denumire}"
        ch_row += 1

    inst_row = 2
    for inst in PnaInstitution.objects.order_by("nume"):
        ws_lists[f"I{inst_row}"] = inst.nume
        inst_row += 1

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_lists.column_dimensions[col].width = 34

    # validări
    max_rows = 500
    header_to_col_main = {header: i for i, (header, _w, _c) in enumerate(_TEMPLATE_PROJECT_COLUMNS, start=1)}
    _add_list_validation(
        ws_main,
        f"{get_column_letter(header_to_col_main['Status implementare'])}2:{get_column_letter(header_to_col_main['Status implementare'])}{max_rows}",
        f"=Liste!$B$2:$B${len(PnaProject.STATUS_IMPLEMENTARE_CHOICES)+1}",
    )
    _add_list_validation(ws_main, f"{get_column_letter(header_to_col_main['Complexitate'])}2:{get_column_letter(header_to_col_main['Complexitate'])}{max_rows}", "=Liste!$C$2:$C$6")
    _add_list_validation(ws_main, f"{get_column_letter(header_to_col_main['Prioritate (1-3)'])}2:{get_column_letter(header_to_col_main['Prioritate (1-3)'])}{max_rows}", "=Liste!$D$2:$D$4")
    _add_list_validation(ws_main, f"{get_column_letter(header_to_col_main['Disponibilitate expertiză internă'])}2:{get_column_letter(header_to_col_main['Disponibilitate expertiză internă'])}{max_rows}", "=Liste!$E$2:$E$4")

    bool_headers = [
        "Necesită expertiză externă",
        "Raport de extindere 2023",
        "Raport de extindere 2024",
        "Raport de extindere 2025",
        "Raport de extindere 2026",
        "Raport de extindere 2027",
        "Planul de creștere economică",
        "Necesită avizare Comisia Europeană",
        "Întârziat 2025",
    ]
    for header in bool_headers:
        col = header_to_col_main[header]
        _add_list_validation(ws_main, f"{get_column_letter(col)}2:{get_column_letter(col)}{max_rows}", "=Liste!$A$2:$A$3")

    header_to_col_acts = {header: i for i, (header, _w, _c) in enumerate(_TEMPLATE_ACT_COLUMNS, start=1)}
    _add_list_validation(ws_acts, f"{get_column_letter(header_to_col_acts['Tip transpunere'])}2:{get_column_letter(header_to_col_acts['Tip transpunere'])}{max_rows}", "=Liste!$F$2:$F$3")

    return wb


def build_pna_import_template_bytes() -> bytes:
    wb = build_pna_import_template_workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
